"""
Servicio para procesar archivos XLSX de PUC y convertirlos a entidades de dominio.
"""
from __future__ import annotations

import logging
from typing import BinaryIO

from app.domain.puc import PUCAccount

logger = logging.getLogger(__name__)

try:
    import openpyxl
    import xlrd
except ModuleNotFoundError:
    openpyxl = None
    xlrd = None


class PUCExcelParserService:
    """
    Servicio para leer archivos Excel (.xlsx y .xls) con estructura de PUC
    y convertirlos a entidades de dominio.
    
    Soporta:
    - .xlsx (Excel 2007+) con openpyxl
    - .xls (Excel antiguo) con xlrd
    """
    
    # Mapeo de nombres de columnas en español a nombres de campos
    COLUMN_MAPPING = {
        "código": "codigo",
        "codigo": "codigo",
        "nombre": "nombre",
        "categoría": "categoria",
        "categoria": "categoria",
        "clase": "clase",
        "relación con": "relacion_con",
        "relacion con": "relacion_con",
        "maneja vencimientos": "maneja_vencimientos",
        "diferencia fiscal": "diferencia_fiscal",
        "activo": "activo",
        "nivel agrupación": "nivel_agrupacion",
        "nivel agrupacion": "nivel_agrupacion",
    }
    
    def __init__(self):
        if openpyxl is None:
            raise RuntimeError("openpyxl no está instalado. Ejecuta: pip install openpyxl xlrd")
    
    def parse_excel(self, file_content: bytes, owner_id: str, filename: str = "") -> list[PUCAccount]:
        """
        Detecta automáticamente el tipo de archivo Excel y lo parsea.
        
        Args:
            file_content: Contenido del archivo en bytes
            owner_id: ID del propietario/empresa
            filename: Nombre del archivo (para detectar extensión)
            
        Returns:
            Lista de entidades PUCAccount parseadas
        """
        # Detectar tipo de archivo por magic bytes
        if file_content.startswith(b'PK\x03\x04') or file_content.startswith(b'PK\x05\x06'):
            # Archivo ZIP (Excel .xlsx)
            logger.info("📊 Detectado formato .xlsx")
            return self.parse_xlsx(file_content, owner_id)
        elif file_content.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
            # Archivo OLE2 (Excel .xls antiguo)
            logger.info("📊 Detectado formato .xls (Excel antiguo)")
            return self.parse_xls(file_content, owner_id)
        else:
            # Intentar por extensión del nombre de archivo
            if filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx'):
                logger.info("📊 Intentando parsear como .xls por extensión")
                return self.parse_xls(file_content, owner_id)
            elif filename.lower().endswith('.xlsx'):
                logger.info("📊 Intentando parsear como .xlsx por extensión")
                return self.parse_xlsx(file_content, owner_id)
            else:
                raise ValueError(
                    "Formato de archivo no reconocido. "
                    "Por favor, asegúrate de que sea un archivo Excel válido (.xlsx o .xls)"
                )
    
    
    def parse_xls(self, file_content: bytes, owner_id: str) -> list[PUCAccount]:
        """
        Lee un archivo .xls (Excel antiguo) usando xlrd.
        """
        if xlrd is None:
            raise ValueError("Soporte para archivos .xls no disponible. Instala: pip install xlrd")
        
        try:
            from io import BytesIO
            workbook = xlrd.open_workbook(file_contents=file_content)
            sheet = workbook.sheet_by_index(0)
            
            if sheet.nrows == 0:
                raise ValueError("El archivo Excel está vacío")
            
            # Buscar la fila de encabezados en las primeras 20 filas
            header_row_idx = None
            headers = []
            
            for row_idx in range(min(20, sheet.nrows)):
                potential_headers = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    potential_headers.append(str(cell_value).strip().lower() if cell_value else "")
                
                # Verificar si esta fila contiene "código" o "codigo"
                if any("codigo" in h or "código" in h for h in potential_headers):
                    header_row_idx = row_idx
                    headers = potential_headers
                    logger.info(f"📋 Encabezados encontrados en fila {row_idx + 1}")
                    break
            
            if header_row_idx is None:
                raise ValueError(
                    "No se encontró la fila de encabezados. "
                    "Asegúrate de que el archivo tenga una fila con al menos las columnas 'Código' y 'Nombre'"
                )
            
            # Mapear índices de columnas
            column_indices = self._map_column_indices(headers)
            
            if "codigo" not in column_indices or "nombre" not in column_indices:
                raise ValueError(
                    "El archivo debe tener al menos las columnas 'Código' y 'Nombre'"
                )
            
            # Leer las filas de datos
            accounts = []
            for row_idx in range(header_row_idx + 1, sheet.nrows):
                try:
                    row_values = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_values.append(cell_value)
                    
                    account = self._parse_row(tuple(row_values), column_indices, owner_id)
                    if account:
                        accounts.append(account)
                except Exception as e:
                    logger.warning(f"⚠️ Error parseando fila {row_idx + 1}: {e}")
                    continue
            
            logger.info(f"✅ Parseadas {len(accounts)} cuentas PUC del archivo .xls")
            return accounts
            
        except Exception as e:
            logger.error(f"❌ Error parseando archivo .xls: {e}")
            raise ValueError(f"Error procesando archivo .xls: {str(e)}") from e
    
    def parse_xlsx(self, file_content: bytes | BinaryIO, owner_id: str) -> list[PUCAccount]:
        """
        Lee un archivo XLSX y retorna una lista de entidades PUCAccount.
        
        Args:
            file_content: Contenido del archivo XLSX en bytes o archivo abierto
            owner_id: ID del propietario/empresa
            
        Returns:
            Lista de entidades PUCAccount parseadas
            
        Raises:
            ValueError: Si el archivo no tiene el formato esperado
        """
        try:
            # Validar que sea un archivo Excel válido verificando el magic number
            if isinstance(file_content, bytes):
                # ZIP magic numbers (Excel es un ZIP)
                if not (file_content.startswith(b'PK\x03\x04') or file_content.startswith(b'PK\x05\x06')):
                    raise ValueError(
                        "El archivo no parece ser un Excel válido. "
                        "Por favor, verifica que el archivo se pueda abrir con Excel y guárdalo nuevamente."
                    )
            
            # Cargar el workbook con más opciones de compatibilidad
            from io import BytesIO
            if isinstance(file_content, bytes):
                file_content = BytesIO(file_content)
            
            workbook = openpyxl.load_workbook(
                file_content, 
                data_only=True,
                read_only=False  # Cambiar a False para mejor compatibilidad
            )
            sheet = workbook.active
            
            if sheet is None:
                raise ValueError("El archivo Excel no tiene hojas")
            
            # Buscar la fila de encabezados en las primeras 20 filas
            header_row_idx = None
            headers = []
            
            for row_idx in range(1, min(21, sheet.max_row + 1)):
                potential_headers = []
                for cell in sheet[row_idx]:
                    if cell.value:
                        potential_headers.append(str(cell.value).strip().lower())
                    else:
                        potential_headers.append("")
                
                # Verificar si esta fila contiene "código" o "codigo"
                if any("codigo" in h or "código" in h for h in potential_headers):
                    header_row_idx = row_idx
                    headers = potential_headers
                    logger.info(f"📋 Encabezados encontrados en fila {row_idx}")
                    break
            
            if header_row_idx is None:
                raise ValueError(
                    "No se encontró la fila de encabezados. "
                    "Asegúrate de que el archivo tenga una fila con al menos las columnas 'Código' y 'Nombre'"
                )
            
            # Mapear índices de columnas
            column_indices = self._map_column_indices(headers)
            
            if "codigo" not in column_indices or "nombre" not in column_indices:
                raise ValueError(
                    "El archivo debe tener al menos las columnas 'Código' y 'Nombre'"
                )
            
            # Leer las filas de datos (desde después de los encabezados)
            accounts = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                try:
                    account = self._parse_row(row, column_indices, owner_id)
                    if account:
                        accounts.append(account)
                except Exception as e:
                    logger.warning(f"⚠️ Error parseando fila {row_idx}: {e}")
                    continue
            
            logger.info(f"✅ Parseadas {len(accounts)} cuentas PUC del archivo Excel")
            return accounts
            
        except ValueError:
            # Re-lanzar ValueError con el mensaje original
            raise
        except Exception as e:
            logger.error(f"❌ Error parseando archivo Excel: {e}")
            error_msg = str(e)
            
            # Mensajes de error más amigables
            if "does not support" in error_msg or "file format" in error_msg:
                raise ValueError(
                    "El archivo no se pudo procesar. Por favor:\n"
                    "1. Abre el archivo con Microsoft Excel o LibreOffice\n"
                    "2. Verifica que se vea correctamente\n"
                    "3. Guárdalo nuevamente como 'Excel Workbook (.xlsx)'\n"
                    "4. Intenta subir el archivo nuevamente"
                ) from e
            elif "Worksheet" in error_msg or "sheet" in error_msg.lower():
                raise ValueError(
                    "El archivo Excel no contiene hojas de trabajo válidas"
                ) from e
            else:
                raise ValueError(f"Error al procesar el archivo: {error_msg}") from e
    
    def _map_column_indices(self, headers: list[str]) -> dict[str, int]:
        """
        Mapea los nombres de columnas a sus índices.
        Retorna un diccionario {nombre_campo: índice_columna}
        """
        indices = {}
        for idx, header in enumerate(headers):
            field_name = self.COLUMN_MAPPING.get(header.strip().lower())
            if field_name:
                indices[field_name] = idx
        
        logger.debug(f"Columnas mapeadas: {indices}")
        return indices
    
    def _parse_row(
        self, 
        row: tuple, 
        column_indices: dict[str, int], 
        owner_id: str
    ) -> PUCAccount | None:
        """
        Parsea una fila del Excel y retorna una entidad PUCAccount.
        Retorna None si la fila está vacía o no es válida.
        """
        # Obtener valores de las columnas
        codigo = self._get_cell_value(row, column_indices.get("codigo"))
        nombre = self._get_cell_value(row, column_indices.get("nombre"))
        
        # Validar campos obligatorios
        if not codigo or not nombre:
            return None
        
        # Si el código no es numérico o alfanumérico, es probablemente un encabezado repetido
        codigo_str = str(codigo).strip()
        if not codigo_str or codigo_str.lower() in ["código", "codigo"]:
            return None
        
        # Crear la entidad
        return PUCAccount.create(
            owner_id=owner_id,
            codigo=codigo_str,
            nombre=str(nombre).strip(),
            categoria=self._get_cell_value(row, column_indices.get("categoria"), ""),
            clase=self._get_cell_value(row, column_indices.get("clase"), ""),
            relacion_con=self._get_cell_value(row, column_indices.get("relacion_con"), ""),
            maneja_vencimientos=self._get_cell_value(row, column_indices.get("maneja_vencimientos"), ""),
            diferencia_fiscal=self._get_cell_value(row, column_indices.get("diferencia_fiscal"), ""),
            activo=self._get_cell_value(row, column_indices.get("activo"), ""),
            nivel_agrupacion=self._get_cell_value(row, column_indices.get("nivel_agrupacion"), ""),
        )
    
    def _get_cell_value(self, row: tuple, column_idx: int | None, default: str = "") -> str:
        """
        Obtiene el valor de una celda de forma segura.
        Retorna default si el índice es None o está fuera de rango.
        """
        if column_idx is None:
            return default
        
        if column_idx >= len(row):
            return default
        
        value = row[column_idx]
        if value is None:
            return default
        
        return str(value).strip()
